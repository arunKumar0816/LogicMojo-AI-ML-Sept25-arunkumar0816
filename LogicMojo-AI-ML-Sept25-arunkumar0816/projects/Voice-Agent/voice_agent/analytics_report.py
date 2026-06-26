"""
analytics_report.py — reads conversations.csv and produces:
  • data/analytics.csv   (daily aggregates)
  • data/report.xlsx     (formatted Excel workbook)

"""

import json
import csv
from pathlib import Path
from collections import Counter

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR  = Path("data")
CONV_CSV  = DATA_DIR / "conversations.csv"
ANAL_CSV  = DATA_DIR / "analytics.csv"
XLSX_OUT  = DATA_DIR / "report.xlsx"


# ── helpers ──────────────────────────────────────────────────────────────────
def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill(color="1F4E79"):
    return PatternFill("solid", start_color=color)


def style_header_row(ws, row: int, ncols: int, color="1F4E79"):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = _header_fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


# ── core ─────────────────────────────────────────────────────────────────────
def load_conversations() -> pd.DataFrame:
    if not CONV_CSV.exists():
        raise FileNotFoundError(f"{CONV_CSV} not found. Run the agent first.")
    df = pd.read_csv(CONV_CSV, parse_dates=["timestamp"])
    df["date"] = df["timestamp"].dt.date
    df["hour"] = df["timestamp"].dt.hour
    return df


def build_daily_agg(df: pd.DataFrame) -> pd.DataFrame:
    return (
        df.groupby("date")
        .agg(
            total_turns=("turn_id", "count"),
            unique_sessions=("session_id", "nunique"),
            avg_response_ms=("response_time_ms", "mean"),
            top_intent=("detected_intent", lambda x: x.mode()[0] if len(x) else "N/A"),
        )
        .reset_index()
        .round({"avg_response_ms": 1})
    )


def build_intent_agg(df: pd.DataFrame) -> pd.DataFrame:
    counts = df["detected_intent"].value_counts().reset_index()
    counts.columns = ["intent", "count"]
    counts["pct"] = (counts["count"] / counts["count"].sum() * 100).round(1)
    return counts


def write_xlsx(df_conv: pd.DataFrame, df_daily: pd.DataFrame, df_intent: pd.DataFrame):
    wb = Workbook()

    # ── Sheet 1: Full conversation log ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Conversation Log"
    cols = [
        "timestamp", "session_id", "turn_id", "user_transcript",
        "detected_intent", "ai_response", "response_time_ms",
    ]
    ws1.append(cols)
    style_header_row(ws1, 1, len(cols), "1F4E79")
    for _, row in df_conv[cols].iterrows():
        ws1.append(list(row))
    for r in ws1.iter_rows(min_row=2):
        for cell in r:
            cell.border    = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws1.freeze_panes = "A2"
    auto_width(ws1)

    # ── Sheet 2: Daily Summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Daily Summary")
    ws2.append(list(df_daily.columns))
    style_header_row(ws2, 1, len(df_daily.columns), "2E75B6")
    for _, row in df_daily.iterrows():
        ws2.append(list(row))
    for r in ws2.iter_rows(min_row=2):
        for cell in r:
            cell.border = _border()
    auto_width(ws2)

    # ── Sheet 3: Intent Breakdown ────────────────────────────────────────────
    ws3 = wb.create_sheet("Intent Breakdown")
    ws3.append(["Intent", "Count", "Percentage (%)"])
    style_header_row(ws3, 1, 3, "375623")
    for _, row in df_intent.iterrows():
        ws3.append([row["intent"], row["count"], row["pct"]])
    # Colour-code by pct
    for r_idx, row in enumerate(ws3.iter_rows(min_row=2), start=2):
        pct = ws3.cell(r_idx, 3).value or 0
        fill_color = "C6EFCE" if pct >= 20 else ("FFEB9C" if pct >= 10 else "FFC7CE")
        for cell in row:
            cell.fill   = PatternFill("solid", start_color=fill_color)
            cell.border = _border()
    auto_width(ws3)

    # ── Sheet 4: KPI Summary ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("KPI Summary")
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 22
    kpis = [
        ("Total Conversations",      len(df_conv)),
        ("Unique Sessions",           df_conv["session_id"].nunique()),
        ("Avg Response Time (ms)",    round(df_conv["response_time_ms"].mean(), 1)),
        ("Max Response Time (ms)",    df_conv["response_time_ms"].max()),
        ("Min Response Time (ms)",    df_conv["response_time_ms"].min()),
        ("Most Common Intent",        df_conv["detected_intent"].mode()[0]),
        ("Days Active",               df_conv["date"].nunique()),
        ("Avg Turns per Session",     round(len(df_conv) / max(df_conv["session_id"].nunique(), 1), 1)),
    ]
    ws4.append(["Metric", "Value"])
    style_header_row(ws4, 1, 2, "7030A0")
    for metric, value in kpis:
        row = ws4.append([metric, value])
    for r in ws4.iter_rows(min_row=2):
        for cell in r:
            cell.border    = _border()
            cell.alignment = Alignment(horizontal="left")

    wb.save(XLSX_OUT)
    print(f"[Report] Saved → {XLSX_OUT}")


def run():
    df = load_conversations()
    if df.empty:
        print("[Report] No conversation data found.")
        return

    daily  = build_daily_agg(df)
    intent = build_intent_agg(df)

    daily.to_csv(ANAL_CSV, index=False)
    print(f"[Report] Daily analytics → {ANAL_CSV}")

    write_xlsx(df, daily, intent)
    print("[Report] Done.")


if __name__ == "__main__":
    run()
